import requests
import csv
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import base64
import json
# Python 3 code to demonstrate the
# working of MD5 (string - hexadecimal)

import hashlib
import random
import string


bvns=[]
api=[]
with open("bvn.csv","r",encoding="utf-8") as f:
    for i in csv.reader(f):
        bvns.append(i[0])
with open("api.csv","r",encoding="utf-8") as f:
    for i in csv.reader(f):
        try:
            api.append(i[0].split(":")[1].strip())
        except:
            pass
       
open_pip_api_key=api[2]
open_pipe_secret_key=api[3]

def get_random_string(length):
    # choose from all lowercase letter
    letters = string.ascii_uppercase + string.ascii_lowercase + string.digits
    result_str = ''.join(random.choice(letters) for i in range(length))
    # print("Random string of length", length, "is:", result_str)
    return result_str
def mdfHash(strToecnrypt):
# initializing string
    result = hashlib.md5(strToecnrypt.encode())
    return result.hexdigest()

def convertImg(imagebase64):
    #open file with base64 string data
    encoded_data = imagebase64
    #decode base64 string data
    decoded_data=base64.b64decode((encoded_data))
    #write the decoded data back to original format in  file
    img_file = open('image.jpeg', 'wb')
    img_file.write(decoded_data)
    img_file.close()
def create_workbook_new(path):
    workbook = Workbook()
    headers="first name,surname,title,sex,Date of birth,BVN,Telephone,state of origin,Local government of origin,state of residence,email,residence address,picture (jpg),bank (fidelity),	Fidelity bank account number	,Supportmfb account number"
    headers=headers.split(",")
    sheet = workbook.active
    sheet.append(headers)
    workbook.save(path)
def create_workbook(path,data,index):
    workbook = load_workbook(path)
   
    sheet = workbook.active
    sheet.append(data)
    image=Image("image.jpeg")
    image.height=80
    image.width=100
    
    
    sheet.add_image(image,f'M{index}')
    workbook.save(path)


def makeBankAccount(bvn_Details,bvn):
    keys="first_name,last_name,title,gender,date_of_birth,bvn,phone_number1,state_of_origin,lga_of_origin,state_of_residence,email,residential_address"
    url = "https://api.onepipe.io/v2/transact"
    req_ref=get_random_string(10)
    trans_Ref=get_random_string(10)
    req_ref_encrypt=mdfHash(str(req_ref)+";"+open_pipe_secret_key)
    payload={
        "request_ref": req_ref,
        "request_type": "open_account",
        "auth": {
        "type": bvn,
        # "secure": "sULdhO/MCgq66D3Bcj4E2Q==",
        "auth_provider": None,
        "route_mode": None
        },
        "transaction": {
        "mock_mode": None,
        "transaction_ref": trans_Ref,
        "transaction_desc": "Opening account",
        "transaction_ref_parent": None,
        "amount": 0,
        "customer": {
        "customer_ref": "234"+bvn_Details['phone_number1'][1:],#customer phone number in 234 format
       
        "firstname": bvn_Details['first_name'],
        "surname": bvn_Details['last_name'],
        "email": bvn_Details['email'],
        "mobile_no": "234"+bvn_Details['phone_number1'][1:]
        },
        "meta": {
        "grb_status": "grb-active"
        },
        "details": {
        "name_on_account": bvn_Details['name_on_card'],
        "middlename": bvn_Details['middle_name'],
        "dob": bvn_Details['date_of_birth'],
        "gender": bvn_Details['gender'][0],
        "title": bvn_Details['title'],
        "address_line_1": None,
        "address_line_2": None,
        "city": None,
        "state": None,
        "country": None
      
        }
        }
        }
   
    print(type(payload))
    payload=json.dumps(payload)
    headers = {
    'Authorization': 'Bearer '+open_pip_api_key,
    'Signature': req_ref_encrypt,
    'Content-Type': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data=payload)

    return response.json()



all_data=[]
create_workbook_new("data.xlsx")
i=2
if api!=[]:
    for bvn in bvns:
        # try:
            url = f"https://api.dojah.io/api/v1/kyc/bvn/full?bvn={bvn}"
            print("getting data of "+bvn)
            headers = {
                "Accept": "text/plain",
                "AppId": api[0],
                "Authorization": api[1]
            }
            response = requests.get(url, headers=headers)
            # print(response.content)
            data= dict(response.json())['entity']
            # print(data)
            # print(data)
            convertImg(data['image'])
            keys1="first_name,last_name,title,gender,date_of_birth,bvn,phone_number1,state_of_origin,lga_of_origin,state_of_residence,email,residential_address"
            
            data_to_append=[]
            print("making bank acoucnt")
            bankAccount=makeBankAccount(data,bvn)
            for key in keys1.split(","):
                data_to_append.append(data[key])
            data_to_append.append("")
            data_to_append.append("fidelity")
            data_to_append.append(bankAccount['data']['provider_response']['account_number'])

            data_to_append.append(bankAccount['data']['provider_response']['meta']['alt_accounts'][0]['account_number'])
            create_workbook("data.xlsx", data_to_append, i)
            
            i+=1
            break

        # except Exception as err:
        #     print("something went wonrg "+ str(err))


else:
    print("Please add apis key in api.csv")



    

